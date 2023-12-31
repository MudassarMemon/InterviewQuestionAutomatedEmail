import smtplib
from dotenv import load_dotenv
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import date
import openpyxl
import os
import traceback 

""" 
Using crontab entry below failed to run script when computer in sleep: 
0 9 * * * /Users/mudassarmemon/.pyenv/shims/python3 /Users/mudassarmemon/Documents/GitHub/InterviewQuestionAutomatedEmail/InterviewQuestionAutomatedEmail.py

I removed the crontab entry and opted to create launch agent to schedule automate script execution --
launchd runs script at 9AM, however, if computer is asleep, script will still run on load.
"""
load_dotenv()

# File logs success/error messages
message_log_file = "/Users/mudassarmemon/Documents/GitHub/InterviewQuestionAutomatedEmail/message_log.txt"

def read_questions_from_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    questions = [[sheet.cell(row=i, column=1).value, sheet.cell(row=i, column=2).value, sheet.cell(row=i, column=3).value] for i in range(4, sheet.max_row + 1)]
    return questions

def send_email(subject, body, to_email, smtp_server, smtp_port, sender_email, sender_password):
    message = MIMEMultipart()
    message['From'] = sender_email
    message['To'] = to_email
    message['Subject'] = subject
    message.attach(MIMEText(body, 'plain'))
    try:
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.sendmail(sender_email, to_email, message.as_string())

        #log success
        with open(message_log_file, "a") as log_file:
            log_file.write("Email sent successfully on " + str(date.today()) + ".\n")

    except Exception as e:
        # Log error message along with the exception details
        with open(message_log_file, "a") as log_file:
            log_file.write(f"Error sending email on {str(date.today())}: {str(e)}\n")
            log_file.write(traceback.format_exc() + "\n")  # Include traceback information

if __name__ == "__main__":
    email_subject = "Coding Interview Question"
    email_body = "Today's daily coding interview question is on {}:\n\n{}\n\nStuck? Find the correct answer here: {}"
    receiver_email = "mudassar95memon@gmail.com"
    smtp_server = "smtp.gmail.com"
    smtp_port = 587
    sender_email = "mudassarsdailycodingquestion@gmail.com"

    sender_password = os.environ.get("APP_PW")

    excel_file_path = "/Users/mudassarmemon/Documents/GitHub/InterviewQuestionAutomatedEmail/InterviewQuestions.xlsx"

    # Keep track of the last sent question
    sent_question_file = "/Users/mudassarmemon/Documents/GitHub/InterviewQuestionAutomatedEmail/sent_question.txt"

    # Read the last sent question
    if os.path.exists(sent_question_file):
        with open(sent_question_file, "r") as file:
            last_sent_question = int(file.read().strip())
    else:
        last_sent_question = 0

    questions = read_questions_from_excel(excel_file_path)

    # Send the next question
    if last_sent_question < len(questions) - 1:
        next_question = questions[last_sent_question][1]
        subject = questions[last_sent_question][0]
        answer = questions[last_sent_question][2]

        email_text = email_body.format(subject, next_question, answer)
        send_email(email_subject, email_text, receiver_email, smtp_server, smtp_port, sender_email, sender_password)

        # Update the last sent question
        with open(sent_question_file, "w") as file:
            file.write(str(last_sent_question + 1))

        print(f"Question sent: {next_question}")
    else:
        print("All questions have been sent.")
